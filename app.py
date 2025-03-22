from flask import Flask, render_template, request, redirect, url_for, flash
import os
import cv2
import numpy as np
from tensorflow.keras.models import load_model, Sequential  # type: ignore
from tensorflow.keras.preprocessing.image import img_to_array  # type: ignore
from tensorflow.keras.layers import Dense, Flatten, Conv2D, MaxPooling2D  # type: ignore
from tensorflow.keras.applications.inception_v3 import InceptionV3  # type: ignore
from tensorflow.keras.layers import Dense, GlobalAveragePooling2D  # type: ignore
from tensorflow.keras.models import Model  # type: ignore
from keras.layers import Input, Conv2D, GlobalAveragePooling2D, Dense, Lambda
from tensorflow.keras.preprocessing.image import ImageDataGenerator  # type: ignore
import pandas as pd
import tensorflow as tf
import matplotlib.pyplot as plt  # type: ignore
import io
import base64
import openpyxl  # type: ignore

app = Flask(__name__)
app.secret_key = "attendance_secret_key"

# Paths
DATASET_DIR = "face_recognition_dataset"
MODEL_PATH = "face_model.h5"
ATTENDANCE_FILE = "attendance1.xlsx"

# Load Class Indices
if os.path.exists("class_new.npy"):
    class_indices = np.load("class_new.npy", allow_pickle=True).item()
else:
    class_indices = {}

reverse_mapping = {v: k for k, v in class_indices.items()}
print(reverse_mapping)
# Define dataset and model paths
DATASET_DIR = r"face_recognition_dataset"
MODEL_PATH = r"model.h5"


if not os.path.exists(ATTENDANCE_FILE):
    wb = openpyxl.Workbook()  # Create a new workbook if file doesn't exist
    sheet = wb.active
    sheet.append(["Name"])  # Add header row for names only
    
    # Add student names from class_indices.values() to the Excel file
    for student_name in reverse_mapping.values():  # Directly using values to get student names
        sheet.append([student_name])  # Add student names without any default status
    
    wb.save(ATTENDANCE_FILE) 


# Load data for training
def load_data(img_height, img_width):
    # Data Augmentation
    datagen = ImageDataGenerator(
        rescale=1.0 / 255,            # Normalizing the image
        validation_split=0.3,         # Split dataset for validation
        rotation_range=30,            # Randomly rotate images by 30 degrees
        width_shift_range=0.2,        # Horizontally shift images by 20% of the width
        height_shift_range=0.2,       # Vertically shift images by 20% of the height
        shear_range=0.2,              # Shear transformation with a shear intensity of 0.2
        zoom_range=0.2,               # Randomly zoom in on images by 20%
        horizontal_flip=True,         # Randomly flip images horizontally
        fill_mode='nearest'   
    )

    train_data = datagen.flow_from_directory(
        DATASET_DIR,
        target_size=(img_height, img_width),
        batch_size=32,
        class_mode="categorical",
        subset="training"
    )
    val_data = datagen.flow_from_directory(
        DATASET_DIR,
        target_size=(img_height, img_width),
        batch_size=32,
        class_mode="categorical",
        subset="validation"
    )

    return train_data, val_data


# Create the model
def create_model(input_shape, num_classes):
    # Create the Base Model
    base_model = InceptionV3(weights='imagenet', include_top=False, input_shape=input_shape)
    
    # Freeze the base model (optional)
    for layer in base_model.layers:
        layer.trainable = False
    
    # Add Custom Top Layers
    x = base_model.output
    x = GlobalAveragePooling2D()(x)
    
    predictions = Dense(num_classes, activation='softmax')(x)  # Adjust the number of classes as needed
    
    # Create the Final Model
    model = Model(inputs=base_model.input, outputs=predictions)
    
    # Compile the Model
    model.compile(optimizer='adam', loss='categorical_crossentropy', metrics=['accuracy'])
    
    model.summary()

    return model

# Train the model
@app.route("/train", methods=["GET", "POST"])
def train():
    if request.method == "POST":
        # Prepare the dataset
        image_data = []
        labels = []
        label_count = 0
        class_indices = {}

        # for folder_name in os.listdir(DATASET_DIR):
        #     folder_path = os.path.join(DATASET_DIR, folder_name)
        #     if os.path.isdir(folder_path):
        #         for img_name in os.listdir(folder_path):
        #             img_path = os.path.join(folder_path, img_name)
        #             img = cv2.imread(img_path)  # Read image in BGR format
        #             img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)  # Convert to RGB
        #             img = cv2.resize(img, (100, 100))  # Resize to match input shape
        #             img = img_to_array(img)
        #             image_data.append(img)
        #             labels.append(label_count)

        #         class_indices[folder_name] = label_count
        #         label_count += 1

        # Save class indices
        

        # image_data = np.array(image_data) / 255.0
        # labels = np.array(labels)

        # # One-hot encode the labels
        # num_classes = len(class_indices)
        # labels = np.eye(num_classes)[labels]
        # print(f"no of classes : {num_classes}")
        # Create and train the model
        train_data, val_data = load_data(100, 100)
        num_classes = len(train_data.class_indices)
        model = create_model((100, 100, 3), num_classes)
        model.fit(train_data, validation_data=val_data, epochs=20)

        # Save the model
        model.save(MODEL_PATH)
        flash("Model trained successfully!", "success")
        return redirect(url_for("home"))

    return render_template("train.html")



@app.route("/")
def home():
    return render_template("home.html")


@app.route("/register", methods=["GET", "POST"])
def register():
    if request.method == "POST":
        user_name = request.form.get("username")
        if not user_name:
            flash("Please provide a user name", "error")
            return redirect(url_for("register"))

        save_dir = os.path.join(DATASET_DIR, user_name)
        if not os.path.exists(save_dir):
            os.makedirs(save_dir)

        cap = cv2.VideoCapture(0)
        count = 0
        while count < 100:  # Capture 100 images
            ret, frame = cap.read()
            if ret:
                face_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + "haarcascade_frontalface_default.xml")
                gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
                faces = face_cascade.detectMultiScale(gray, 1.3, 5)

                for (x, y, w, h) in faces:
                    # Crop the grayscale face
                    face_img_gray = gray[y:y+h, x:x+w]
                    img_path = os.path.join(save_dir, f"{user_name}_{count}.jpg")
                    # Save the grayscale image
                    cv2.imwrite(img_path, face_img_gray)
                    count += 1

                cv2.imshow("Registering User", gray)  # Display the grayscale frame

            if cv2.waitKey(1) & 0xFF == ord('q') or count >= 100:
                break

        cap.release()
        cv2.destroyAllWindows()
        flash(f"User {user_name} registered successfully!", "success")
        return redirect(url_for("home"))

    return render_template("register.html")



# from datetime import datetime

# @app.route("/mark_attendance1", methods=["GET", "POST"])
# def mark_attendance1():
#     attendance = {}  # Store attendance names

#     if request.method == "POST":
#         # Check if the model is available
#         if not os.path.exists(MODEL_PATH):
#             flash("Model not found! Train the CNN first.", "error")
#             return redirect(url_for("home"))

#         # Load the pre-trained CNN model
#         model = load_model(MODEL_PATH)

#         # Get the uploaded image from the form
#         file = request.files.get("image")
#         if not file:
#             flash("No image uploaded!", "error")
#             return redirect(url_for("mark_attendance"))

#         # Save the uploaded image temporarily
#         img_path = os.path.join("uploads", file.filename)
#         file.save(img_path)

#         # Load the image
#         img = cv2.imread(img_path)
#         face_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + "haarcascade_frontalface_default.xml")
#         gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
#         faces = face_cascade.detectMultiScale(gray, 1.3, 5)

#         # Iterate over each face detected
#         for (x, y, w, h) in faces:
#             # Preprocess the face for the model
#             face_img = img[y:y+h, x:x+w]
#             face_img = cv2.cvtColor(face_img, cv2.COLOR_BGR2RGB)  # Convert to RGB
#             face_img = cv2.resize(face_img, (100, 100))           # Resize to model's input size
#             face_img = face_img / 255.0                           # Normalize
#             face_img = np.expand_dims(face_img, axis=0)           # Add batch dimension

#             # Make a prediction
#             prediction = model.predict(face_img)
#             predicted_class = np.argmax(prediction)
#             class_name = reverse_mapping.get(predicted_class, "Unknown")
#             attendance[class_name] = "Present"

#             # Draw a rectangle and label on the image
#             cv2.rectangle(img, (x, y), (x+w, y+h), (255, 0, 0), 2)
#             cv2.putText(img, class_name, (x, y-10), cv2.FONT_HERSHEY_SIMPLEX, 1, (255, 0, 0), 2)

#         # Save the image with annotations (Optional)
#         marked_image_path = os.path.join("uploads", "marked_" + file.filename)
#         cv2.imwrite(marked_image_path, img)

#         # Save attendance to Excel
#         if not os.path.exists(ATTENDANCE_FILE):
#             wb = openpyxl.Workbook()
#             sheet = wb.active
#             sheet.append(["Name", "Status"])  # Add header row
#             wb.save(ATTENDANCE_FILE)

#         wb = openpyxl.load_workbook(ATTENDANCE_FILE)
#         sheet = wb.active
#         current_date = datetime.now().strftime("%Y-%m-%d")  # Format as "YYYY-MM-DD"

#         # Check if the column for the current date exists
#         header = [cell.value for cell in sheet[1]]  # Get the first row (header)
        
#         if current_date not in header:
#             header.append(current_date)  # Add current date to the header list
#             # Re-write the updated header back to the sheet
#             for col_idx, value in enumerate(header, start=1):  # `start=1` to begin from column 1
#                 sheet.cell(row=1, column=col_idx, value=value)

#         # Find the index of the current date in the header
#         date_column_idx = header.index(current_date) + 1  # +1 to convert zero-based index to Excel's column index

#         # Iterate over each row (students)
#         for row in sheet.iter_rows(min_row=2, max_col=sheet.max_column):
#             name = row[0].value  # First column should be "Name"
#             if name in attendance:
#                 try:
#                     row[date_column_idx].value = attendance[name]  # Mark attendance in the corresponding date column
#                 except IndexError:
#                     # In case there is an issue with row access, log the error
#                     flash(f"Error marking attendance for {name}.", "error")

#         wb.save(ATTENDANCE_FILE)

#         flash("Attendance marked successfully!", "success")

#         # Pass the marked attendance names to the template
#         return render_template("mark_attendance.html", attendance=attendance)

#     return render_template("mark_attendance.html")




from datetime import datetime
import cv2
import numpy as np
import openpyxl
from tensorflow.keras.models import load_model # type: ignore
from flask import request, flash, redirect, url_for, render_template
import os

@app.route("/mark_attendance", methods=["GET", "POST"])
def mark_attendance():
    attendance = {}  # Store attendance names

    if request.method == "POST":
        # Check if the model is available
        if not os.path.exists(MODEL_PATH):
            flash("Model not found! Train the CNN first.", "error")
            return redirect(url_for("home"))

        # Load the pre-trained CNN model
        model = load_model(MODEL_PATH)

        # Get the uploaded image from the form
        file = request.files.get("image")
        if not file:
            flash("No image uploaded!", "error")
            return redirect(url_for("mark_attendance"))

        # Save the uploaded image temporarily
        img_path = os.path.join("uploads", file.filename)
        file.save(img_path)

        # Load the image
        img = cv2.imread(img_path)
        face_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + "haarcascade_frontalface_default.xml")
        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        faces = face_cascade.detectMultiScale(gray, 1.3, 5)

        # Iterate over each face detected
        for (x, y, w, h) in faces:
            # Preprocess the face for the model
            face_img = img[y:y+h, x:x+w]
            face_img = cv2.cvtColor(face_img, cv2.COLOR_BGR2RGB)  # Convert to RGB
            face_img = cv2.resize(face_img, (100, 100))           # Resize to model's input size
            face_img = face_img / 255.0                           # Normalize
            face_img = np.expand_dims(face_img, axis=0)           # Add batch dimension

            # Make a prediction
            prediction = model.predict(face_img)
            predicted_class = np.argmax(prediction)
            class_name = reverse_mapping.get(predicted_class, "Unknown")
            attendance[class_name] = "Present"

            # Draw a rectangle and label on the image
            cv2.rectangle(img, (x, y), (x+w, y+h), (255, 0, 0), 2)
            cv2.putText(img, class_name, (x, y-10), cv2.FONT_HERSHEY_SIMPLEX, 1, (255, 0, 0), 2)

        # Save the image with annotations (Optional)
        marked_image_path = os.path.join("uploads", "marked_" + file.filename)
        cv2.imwrite(marked_image_path, img)

        # Save attendance to Excel
        

        wb = openpyxl.load_workbook(ATTENDANCE_FILE)
        sheet = wb.active
        current_date = datetime.now().strftime("%Y-%m-%d")  # Format as "YYYY-MM-DD"

        # Check if the column for the current date exists
        header = [cell.value for cell in sheet[1]]  # Get the first row (header)
        
        if current_date not in header:
            header.append(current_date)  # Add current date to the header list
            # Re-write the updated header back to the sheet
            for col_idx, value in enumerate(header, start=1):  # `start=1` to begin from column 1
                sheet.cell(row=1, column=col_idx, value=value)

        # Find the index of the current date in the header
        date_column_idx = header.index(current_date) + 1  # +1 to convert zero-based index to Excel's column index

        # Iterate over each row (students)
        for row in sheet.iter_rows(min_row=2, max_col=sheet.max_column):
            name = row[0].value  # First column should be "Name"
            if name in attendance:
                try:
                    # Ensure the row has enough columns to write the attendance
                    if len(row) < date_column_idx:
                        for i in range(len(row), date_column_idx):
                            sheet.cell(row=row[0].row, column=i + 1, value="")
                    row[date_column_idx - 1].value = attendance[name]  # Mark attendance in the corresponding date column
                except IndexError:
                    # In case there is an issue with row access, log the error
                    flash(f"Error marking attendance for {name}.", "error")

        wb.save(ATTENDANCE_FILE)

        flash("Attendance marked successfully!", "success")

        # Pass the marked attendance names to the template
        return render_template("mark_attendance.html", attendance=attendance)

    return render_template("mark_attendance.html")

import re
import pandas as pd
from datetime import datetime

import re
import pandas as pd
from datetime import datetime
import io
import base64
import matplotlib.pyplot as plt
from flask import render_template, flash, redirect, url_for

@app.route("/analysis1")
def analysis1():
    if not os.path.exists(ATTENDANCE_FILE):
        flash("No attendance data found!", "error")
        return redirect(url_for("home"))

    # Read the attendance data without specifying column names
    df = pd.read_excel(ATTENDANCE_FILE)

    # Ensure the dataframe has the correct columns: "User Name" and "Status"
    if "Name" not in df.columns:
        flash("Attendance data is malformed!", "error")
        return redirect(url_for("home"))

    # Convert all column names to strings and check for date columns
    df.columns = df.columns.astype(str)
    date_pattern = r'^\d{4}-\d{2}-\d{2}$'
    date_columns = [col for col in df.columns if re.match(date_pattern, col)]

    if not date_columns:
        flash("No date columns found in the attendance data.", "error")
        return redirect(url_for("home"))

    # Group by user and count status for each date column
    user_summary = {}
    for date_column in date_columns:
        # Ensure that we count the "Present" status correctly, even with case sensitivity
        daily_summary = df.groupby("Name")[date_column].apply(lambda x: x.str.contains('Present', case=False).sum())
        user_summary[date_column] = daily_summary

    # Plot attendance summary for each date column
    for date_column, daily_summary in user_summary.items():
        plt.figure(figsize=(10, 5))
        daily_summary.plot(kind="bar", color="skyblue")
        plt.title(f"Attendance Summary for {date_column}")
        plt.xlabel("User Name")
        plt.ylabel("Days Present")
        plt.tight_layout()

        img = io.BytesIO()
        plt.savefig(img, format="png")
        img.seek(0)
        plot_url = base64.b64encode(img.getvalue()).decode()
        plt.close()

        # Render analysis page with the attendance plot for each date
        return render_template("analysis.html", plot_url=plot_url, date_column=date_column)

    flash("No data to display.", "error")
    return redirect(url_for("home"))


@app.route("/analysis")
def analysis():
    if not os.path.exists(ATTENDANCE_FILE):
        flash("No attendance data found!", "error")
        return redirect(url_for("home"))

    # Read the attendance data
    df = pd.read_excel(ATTENDANCE_FILE)

    # Ensure the dataframe has the correct columns: "Name" and date columns
    if "Name" not in df.columns:
        flash("Attendance data is malformed!", "error")
        return redirect(url_for("home"))

    # Convert all column names to strings and check for date columns
    df.columns = df.columns.astype(str)
    date_pattern = r'^\d{4}-\d{2}-\d{2}$'
    date_columns = [col for col in df.columns if re.match(date_pattern, col)]
    print(f"++++++++{date_columns}")

    if not date_columns:
        flash("No date columns found in the attendance data.", "error")
        return redirect(url_for("home"))

    # Calculate the number of days being analyzed
    num_days = len(date_columns)
    flash(f"Analyzing attendance for {num_days} days.", "info")

    # Prepare a list of plots for all date columns
    plots = []
    for date_column in date_columns:
        # Group by user and count "Present" status for each date column
        daily_summary = df.groupby("Name")[date_column].apply(
            lambda x: x.str.contains('Present', case=False).sum() if x.dtype == 'object' else 0
        )

        # Plot attendance summary for the current date
        plt.figure(figsize=(10, 5))
        daily_summary.plot(kind="bar", color="skyblue")
        plt.title(f"Attendance Summary for {date_column}")
        plt.xlabel("User Name")
        plt.ylabel("Days Present")
        plt.tight_layout()

        # Save the plot as a base64-encoded string
        img = io.BytesIO()
        plt.savefig(img, format="png")
        img.seek(0)
        plot_url = base64.b64encode(img.getvalue()).decode()
        plt.close()

        # Append the plot URL and the date column to the list
        plots.append({"plot_url": plot_url, "date_column": date_column})

    # Render the analysis page with all plots
    return render_template("analysis.html", plots=plots, num_days=num_days)



if __name__ == "__main__":
    app.run(threaded=False)
